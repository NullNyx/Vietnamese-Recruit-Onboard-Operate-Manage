"""Excel parser for bulk employee import.

Parses .xlsx files into structured dictionaries, validating required fields
and normalizing date formats.
"""

import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

# Required fields that must be present and non-empty in each row.
REQUIRED_FIELDS = {"full_name", "email"}

# All recognized column headers (snake_case normalized).
KNOWN_FIELDS = {
    "full_name",
    "email",
    "phone",
    "date_of_birth",
    "gender",
    "address",
    "department_name",
    "position_name",
    "start_date",
    "id_number",
    "tax_code",
    "contract_type",
}

# Basic email regex for format validation.
_EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")


def _normalize_header(header: str) -> str:
    """Normalize a column header to snake_case.

    Strips whitespace, lowercases, and replaces spaces/hyphens with underscores.
    """
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def _parse_date(value: object) -> date | None:
    """Parse a date value from an Excel cell.

    Supports:
    - Native Python datetime/date objects (from Excel date cells)
    - String in YYYY-MM-DD format
    - String in DD/MM/YYYY format

    Returns None if the value is empty/None.
    Raises ValueError if the value cannot be parsed.
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    str_value = str(value).strip()
    if not str_value:
        return None

    # Try YYYY-MM-DD
    try:
        return datetime.strptime(str_value, "%Y-%m-%d").date()
    except ValueError:
        pass

    # Try DD/MM/YYYY
    try:
        return datetime.strptime(str_value, "%d/%m/%Y").date()
    except ValueError:
        pass

    raise ValueError(f"Invalid date format: '{str_value}'. Expected YYYY-MM-DD or DD/MM/YYYY")


def _validate_email(email: str) -> bool:
    """Validate email format using a basic regex check."""
    return bool(_EMAIL_REGEX.match(email.strip()))


def parse_excel(file_bytes: bytes) -> tuple[list[dict], list[dict]]:
    """Parse an Excel (.xlsx) file into a list of row dictionaries.

    Reads the first worksheet, uses the first row as headers, and parses
    subsequent rows into dictionaries. Validates required fields and email
    format per row.

    Args:
        file_bytes: Raw bytes of the .xlsx file.

    Returns:
        A tuple of (parsed_rows, errors) where:
        - parsed_rows: List of dicts with successfully parsed row data.
          Date fields are converted to date objects.
        - errors: List of dicts with 'row' (1-indexed Excel row number)
          and 'message' keys describing validation failures.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return [], []

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    # Parse headers from the first row.
    raw_headers = rows[0]
    headers: list[str | None] = []
    for h in raw_headers:
        if h is None:
            headers.append(None)
        else:
            headers.append(_normalize_header(str(h)))

    parsed_rows: list[dict] = []
    errors: list[dict] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        row_data: dict = {}
        row_errors: list[str] = []

        for col_idx, cell_value in enumerate(row):
            if col_idx >= len(headers):
                break
            header = headers[col_idx]
            if header is None or header not in KNOWN_FIELDS:
                continue
            row_data[header] = cell_value

        # Skip entirely empty rows.
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_data.values()):
            continue

        # Validate required fields.
        for field in REQUIRED_FIELDS:
            value = row_data.get(field)
            if value is None or (isinstance(value, str) and not value.strip()):
                row_errors.append(f"Missing required field: {field}")

        # Validate email format.
        email_value = row_data.get("email")
        if email_value is not None and isinstance(email_value, str) and email_value.strip():
            if not _validate_email(email_value):
                row_errors.append(f"Invalid email format: '{email_value.strip()}'")

        # Parse date fields.
        for date_field in ("date_of_birth", "start_date"):
            if date_field in row_data and row_data[date_field] is not None:
                try:
                    row_data[date_field] = _parse_date(row_data[date_field])
                except ValueError as e:
                    row_errors.append(str(e))

        if row_errors:
            for msg in row_errors:
                errors.append({"row": row_idx, "message": msg})
        else:
            # Clean up string values (strip whitespace).
            for key, value in row_data.items():
                if isinstance(value, str):
                    row_data[key] = value.strip()
            # Include the row number for downstream error reporting.
            row_data["_row_number"] = row_idx
            parsed_rows.append(row_data)

    return parsed_rows, errors
